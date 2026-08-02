#!/usr/bin/env python3
"""
Xuất báo cáo ra tệp bảng tính (YC-AN-10, YC-NK-06, YC-DB-08).

HAI ĐIỀU DỄ SAI NHẤT KHI XUẤT DỮ LIỆU TIẾNG VIỆT — cả hai đều được xử lý ở đây:

1. **CSV UTF-8 không có BOM thì Excel trên Windows hiển thị sai dấu.** Excel đoán bảng mã theo
   codepage hệ thống (cp1258/cp1252 ở máy Việt Nam) chứ không mặc định UTF-8. Không có BOM thì
   "Báo cáo tổng kết" thành "BÃ¡o cÃ¡o..." — người dùng báo "xuất ra bị lỗi font" và không ai biết
   sửa ở đâu. Ghi BOM `\\ufeff` là cách duy nhất khiến Excel nhận đúng.

2. **Số điện thoại / mã số bị Excel đổi thành số khoa học.** Mã tài liệu dạng `1234567890123` bị
   hiển thị thành `1,23457E+12`. Với XLSX ta ghi kiểu chuỗi tường minh.

`openpyxl` được import LAZY và có phương án lùi CSV: gói này không có trên máy dev, và trên máy chủ
air-gapped thì nó là một thứ nữa phải tải trước khi ngắt mạng. Chức năng xuất phải chạy được trong
cả hai trường hợp — mất định dạng còn hơn mất chức năng.
"""

import csv
import io
import logging
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Sequence, Tuple

logger = logging.getLogger("core.export")

# BOM UTF-8 — bắt buộc để Excel trên Windows đọc đúng tiếng Việt (xem docstring)
UTF8_BOM = "﻿"


def format_cell(value: Any) -> str:
    """
    Định dạng một ô theo quy ước dự án.

    - Ngày: `DD/MM/YYYY` (lưu `YYYY-MM-DD`, hiển thị `DD/MM/YYYY`)
    - `None`: chuỗi rỗng, KHÔNG phải "None" — "None" trong ô Excel là thứ người dùng không hiểu
    - `True/False`: "Có"/"Không" — bảng báo cáo cho cán bộ, không phải cho lập trình viên
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Có" if value else "Không"
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return str(value)


def build_rows(data: List[Dict], columns: Sequence[Tuple[str, str]]) -> List[List[str]]:
    """
    Chuyển danh sách dict thành ma trận chuỗi, kèm dòng tiêu đề.

    `columns` là danh sách `(khóa, nhãn_tiếng_Việt)` — thứ tự cột do nơi gọi quyết định, và nhãn
    luôn bằng tiếng Việt vì người mở tệp là cán bộ nghiệp vụ, không phải người viết mã.
    """
    rows = [[label for _, label in columns]]
    for item in data:
        rows.append([format_cell(item.get(key)) for key, _ in columns])
    return rows


def to_csv_bytes(rows: List[List[str]]) -> bytes:
    """
    Xuất CSV UTF-8 **có BOM** để Excel trên Windows hiển thị đúng tiếng Việt.

    Dùng `\\r\\n` theo đúng quy ước CSV của Excel — thiếu nó thì một số phiên bản gộp các dòng lại.
    """
    buffer = io.StringIO()
    buffer.write(UTF8_BOM)
    writer = csv.writer(buffer, lineterminator="\r\n")
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8")


def to_xlsx_bytes(sheets: Dict[str, List[List[str]]]) -> Optional[bytes]:
    """
    Xuất XLSX nhiều sheet. Trả `None` nếu `openpyxl` không có — nơi gọi lùi về CSV.

    Mọi ô ghi kiểu CHUỖI tường minh: để Excel tự đoán kiểu sẽ biến mã tài liệu dài thành số khoa học
    (`1,23457E+12`) và cắt số 0 đứng đầu.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.info("Không có openpyxl — xuất CSV thay cho XLSX")
        return None

    workbook = Workbook()
    workbook.remove(workbook.active)      # bỏ sheet mặc định rỗng

    for name, rows in sheets.items():
        # Tên sheet Excel tối đa 31 ký tự và không nhận một số ký tự đặc biệt
        sheet = workbook.create_sheet(_safe_sheet_name(name))
        for row in rows:
            sheet.append(row)

        if rows:
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            sheet.freeze_panes = "A2"     # giữ tiêu đề khi cuộn — bảng báo cáo thường rất dài
            _autosize(sheet, rows, get_column_letter)

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _safe_sheet_name(name: str) -> str:
    for char in "[]:*?/\\":
        name = name.replace(char, "-")
    return name[:31] or "Sheet1"


def _autosize(sheet, rows: List[List[str]], get_column_letter) -> None:
    """Giãn cột theo nội dung, có trần: một ô tóm tắt dài không nên làm cột rộng cả màn hình."""
    widths: Dict[int, int] = {}
    for row in rows:
        for index, value in enumerate(row, start=1):
            widths[index] = min(max(widths.get(index, 10), len(str(value)) + 2), 60)
    for index, width in widths.items():
        sheet.column_dimensions[get_column_letter(index)].width = width


def export(sheets: Dict[str, List[List[str]]]) -> Tuple[bytes, str, str]:
    """
    Xuất bảng tính, tự chọn định dạng khả dụng.

    Trả về `(nội_dung, phần_mở_rộng, kiểu_MIME)`. Khi lùi về CSV chỉ xuất được sheet ĐẦU TIÊN —
    CSV không có khái niệm nhiều sheet; nơi gọi nên đặt bảng chính lên đầu.
    """
    content = to_xlsx_bytes(sheets)
    if content is not None:
        return (content, "xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    first_rows = next(iter(sheets.values()), [])
    return to_csv_bytes(first_rows), "csv", "text/csv; charset=utf-8"
